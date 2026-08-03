"""产出层：生成 Excel + CSV 表格。"""

from __future__ import annotations

import csv
from pathlib import Path

from ..collect.base import Article

HEADERS = ["日期", "账号", "标题", "链接", "摘要", "标签", "正文", "来源"]


def articles_to_rows(articles: list[Article]) -> list[list[str]]:
    """将 Article 列表转换为表格行。"""
    rows: list[list[str]] = []
    for a in articles:
        rows.append(
            [
                a.published_at,
                a.account,
                a.title,
                a.url,
                a.summary,
                "、".join(a.tags),
                a.content,
                a.source,
            ]
        )
    return rows


def write_csv(articles: list[Article], path: str | Path) -> Path:
    """写入 UTF-8 with BOM 的 CSV（Excel 可直接打开中文不乱码）。"""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(articles_to_rows(articles))
    return path


def write_xlsx(articles: list[Article], path: str | Path) -> Path:
    """写入 Excel。依赖 openpyxl。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "资讯"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in articles_to_rows(articles):
        ws.append(row)

    widths = [12, 24, 40, 50, 60, 30, 80, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # 标题/摘要/正文自动换行，方便阅读
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)
    return path
